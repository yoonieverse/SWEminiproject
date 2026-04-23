import time
import math
import re
import requests
import unicodedata
from difflib import SequenceMatcher
from dotenv import load_dotenv
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

load_dotenv()

LIVE_MODE = True
API_KEY = os.getenv("GEMINI_API_KEY")
WIKIPEDIA_API = "https://en.wikipedia.org/w/api.php"
HEADERS = {"User-Agent": "BookVerifier/1.0 (educational project)"}

GEMINI_MAX_CALLS = 20
_gemini_call_count = 0
_gemini_cache = {}

_gemini_client = None
if API_KEY:
    try:
        from google import genai
        _gemini_client = genai.Client(api_key=API_KEY)
    except Exception as _e:
        print(f"[GEMINI INIT] SDK unavailable: {_e}")


# NORMALIZERS
def normalize_title(title: str) -> str:
    """Unicode-safe: keeps non-ASCII glyphs. Used for duplicate detection."""
    nfkd = unicodedata.normalize("NFKD", title)
    return re.sub(r"\s+", " ", nfkd).strip().lower()

def safe_ascii_title(title: str) -> str:
    """Strips to plain ASCII after NFKD decomposition. Used for Wikipedia search queries."""
    nfkd = unicodedata.normalize("NFKD", title)
    ascii_only = nfkd.encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"[^\w\s\-']", " ", ascii_only)
    return re.sub(r"\s+", " ", cleaned).strip().lower()


# TRANSLATOR
def translate_title(original: str) -> tuple[str | None, str]:
    """Translates any title to English via deep-translator (source=auto). Returns (result, status)."""
    try:
        from deep_translator import GoogleTranslator
        result = GoogleTranslator(source="auto", target="en").translate(original)
        if not result:
            return None, "empty_result"
        result = result.strip()
        if result.lower() == original.strip().lower():
            return None, "same_text"
        return result, "translated"
    except Exception as exc:
        return None, f"error:{exc}"


# GEMINI
def _call_gemini(prompt: str) -> str | None:
    """Calls Gemini (any model that will let me abuse their api call rate) with caching and a hard call-count limit."""
    global _gemini_call_count
    if not _gemini_client:
        return None
    cache_key = prompt[:300]
    if cache_key in _gemini_cache:
        return _gemini_cache[cache_key]
    if _gemini_call_count >= GEMINI_MAX_CALLS:
        print(f"  [GEMINI] Hard limit ({GEMINI_MAX_CALLS}) reached -- skipping.")
        return None
    try:
        response = _gemini_client.models.generate_content(
            model="gemini-flash-latest",
            contents=prompt,
        )
        text = response.text.strip()
        _gemini_call_count += 1
        _gemini_cache[cache_key] = text
        return text
    except Exception as exc:
        print(f"  [GEMINI] Error: {exc}")
        return None


# SIMILARITY
def _enhanced_similarity(a: str, b: str) -> float:
    """Case-insensitive SequenceMatcher between two strings."""
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def _title_similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def _method_confidence(method: str, sim: float) -> float:
    """Maps resolution method + similarity score to a calibrated confidence value."""
    bands = {
        "direct":         (0.90, 1.00),
        "retry_original": (0.75, 0.90),
        "failed":         (0.00, 0.00),
    }
    for key, (lo, hi) in bands.items():
        if key in method:
            return round(lo + (hi - lo) * min(sim, 1.0), 3)
    return round(min(sim, 1.0), 3)


# DUPLICATE DETECTION
def _local_duplicate_prefilter(results: list[dict]) -> list[dict]:
    """
    Compares translated titles when available, falling back to unicode-normalized originals.
    """
    assigned: dict[int, str] = {}
    next_group = [1]

    def _get_or_make(num: int) -> str:
        if num not in assigned:
            assigned[num] = f"L{next_group[0]}"
            next_group[0] += 1
        return assigned[num]

    for i, ra in enumerate(results):
        for j in range(i + 1, len(results)):
            rb = results[j]
            a_primary = ra.get("translated_title") or ra["normalized_title"]
            b_primary = rb.get("translated_title") or rb["normalized_title"]
            score = max(
                _enhanced_similarity(a_primary, b_primary),
                _enhanced_similarity(ra["title"], rb["title"]),
            )
            if score > 0.72:
                grp = (assigned.get(ra["num"])
                       or assigned.get(rb["num"])
                       or _get_or_make(ra["num"]))
                assigned[ra["num"]] = grp
                assigned[rb["num"]] = grp

    for r in results:
        if r["num"] in assigned:
            r["duplicate_group"] = assigned[r["num"]]

    return results


def gemini_find_duplicates(results: list[dict]) -> list[dict]:
    """
    Two-stage duplicate detection:
      Stage 1 - local pre-filter (fast, free)
      Stage 2 - Gemini batch call for cross-language / semantic duplicates
    Local groups are labelled L1, L2 ...; Gemini groups are labelled G1, G2 ...
    Should expect only Local group outputs due to API rate limit (avoid stage 2)
    """
    if not results:
        return results

    results = _local_duplicate_prefilter(results)

    enriched_lines = []
    for r in results:
        parts = [r.get("title", "")]
        if r.get("translated_title"):
            parts.append(r["translated_title"])
        if r.get("normalized_title") and r["normalized_title"] not in parts:
            parts.append(r["normalized_title"])
        enriched_lines.append(f"{r['num']}. {' | '.join(parts)}")

    prompt = (
        "Below is a numbered list of books. Each entry may show: "
        "original title | translated title | normalized title.\n\n"
        "Identify any entries that refer to the SAME book "
        "(cross-language equivalents, variant editions, different casings).\n"
        "Reply ONLY with a valid JSON array of groups, each group a list of entry numbers. "
        "Example: [[1,5],[3,7]]. If no duplicates exist reply with: []\n\n"
        + "\n".join(enriched_lines)
    )

    raw = _call_gemini(prompt)
    if not raw:
        return results

    try:
        import json
        clean = re.sub(r"```[a-z]*", "", raw).strip().strip("`").strip()
        groups = json.loads(clean)
    except Exception:
        return results

    next_gid = 1
    for members in groups:
        label = f"G{next_gid}"
        next_gid += 1
        for num in members:
            for r in results:
                if r["num"] == num and not r.get("duplicate_group"):
                    r["duplicate_group"] = label

    return results


def gemini_confidence_commentary(stats: dict) -> str:
    prompt = (
        f"A book verification system checked {stats['total']} books against Wikipedia. "
        f"{stats['verified']} were verified ({stats['proportion']*100:.1f}%). "
        f"The 95% confidence interval is {stats['lower']*100:.1f}%--{stats['upper']*100:.1f}%. "
        "Write ONE plain-English sentence interpreting these results for a non-technical reader. "
        "No bullet points, no markdown."
    )
    return _call_gemini(prompt) or ""


# WIKIPEDIA
def _wiki_search(query: str, limit: int = 5) -> list[dict]:
    """Runs a Wikipedia API search and returns the raw result list."""
    params = {
        "action": "query",
        "list": "search",
        "srsearch": f'"{query}" book',
        "srlimit": limit,
        "format": "json",
    }
    try:
        resp = requests.get(WIKIPEDIA_API, params=params, headers=HEADERS, timeout=8)
        return resp.json().get("query", {}).get("search", [])
    except Exception:
        return []


def check_wikipedia(title: str) -> tuple[bool, str, str, float, str]:
    """Searches Wikipedia for a title and returns (found, wiki_title, url, confidence, method)."""
    results = _wiki_search(title)
    if not results:
        return False, "Not found on Wikipedia", "", 0.0, "failed"
    scored = [(r["title"], _title_similarity(title, r["title"])) for r in results]
    best_title, best_sim = max(scored, key=lambda x: x[1])
    if best_sim >= 0.65:
        url = f"https://en.wikipedia.org/wiki/{best_title.replace(' ', '_')}"
        return True, best_title, url, _method_confidence("direct", best_sim), "direct"
    return False, best_title, "", _method_confidence("failed", best_sim), "failed"


# RESOLUTION PIPELINE
def resolve_title(book: dict) -> dict:
    """
    Per-book pipeline:
      1. Translate title (source=auto; translator decides if translation is needed)
      2. Build ASCII search title from translated (or original if untranslated)
      3. Wikipedia check on the ASCII title
      4. If not found, retry with the ASCII form of the original title
    """
    original_title: str = book["title"]

    translated: str | None = None
    translation_status = "skipped"

    translated, translation_status = translate_title(original_title)
    if translated:
        print(f"  [TRANSLATE] '{original_title}' -> '{translated}'  ({translation_status})")
    elif translation_status not in ("skipped", "same_text"):
        print(f"  [TRANSLATE] Could not translate '{original_title}' -- {translation_status}")

    source_for_search = translated if translated else original_title
    normalized = safe_ascii_title(source_for_search)
    unicode_normalized = normalize_title(source_for_search)

    found, wiki_title, url, confidence, method = check_wikipedia(normalized)

    normalized_original = safe_ascii_title(original_title)
    if not normalized_original.strip():
        normalized_original = original_title

    if not found and translated and normalized != normalized_original:
        print(f"  [RETRY] Trying normalized original: '{normalized_original}'")
        found, wiki_title, url, confidence, method = check_wikipedia(normalized_original)
        if found:
            method = f"retry_original/{method}"

    return {
        **book,
        "found": found,
        "translated_title": translated or "",
        "translation_status": translation_status,
        "normalized_title": unicode_normalized,
        "search_title_used": normalized,
        "wiki_title": wiki_title,
        "url": url if found else "N/A",
        "confidence": confidence,
        "resolution_method": method,
        "duplicate_group": "",
    }


# STATS
def compute_confidence_interval(results: list[dict]) -> dict:
    """Computes verification proportion with a 95% confidence interval."""
    total = len(results)
    verified = sum(1 for r in results if r["found"])
    if total == 0:
        return {k: 0 for k in
                ["total", "verified", "not_verified", "proportion",
                 "margin_error", "lower", "upper", "avg_confidence"]}
    p = verified / total
    z = 1.96
    margin_error = z * math.sqrt((p * (1 - p)) / total)
    avg_conf = (
        sum(r["confidence"] for r in results if r["found"]) / verified
        if verified else 0.0
    )
    return {
        "total": total,
        "verified": verified,
        "not_verified": total - verified,
        "proportion": p,
        "margin_error": margin_error,
        "lower": max(0, p - margin_error),
        "upper": min(1, p + margin_error),
        "avg_confidence": round(avg_conf, 3),
    }


# INPUT
def load_books() -> list[dict]:
    """Loads books from books_input.xlsx (columns: num, title, author, genre)."""
    wb = load_workbook("books_input.xlsx")
    ws = wb.active
    books = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[1]:
            continue
        books.append({
            "num": row[0],
            "title": row[1],
            "author": row[2],
            "genre": row[3],
        })
    return books


# TERMINAL OUTPUT
_COL  = 75
_W_ST = 11
_W_NUM = 4
_W_TTL = 46

def print_header():
    print("\n" + "=" * _COL)
    print("BOOK VERIFICATION")
    print("=" * _COL)
    print(
        f"{'STATUS':<{_W_ST}} | "
        f"{'#':<{_W_NUM}} | "
        f"TITLE"
    )
    print("-" * _COL)
    print()

def print_result(result: dict):
    status = "[VERIFIED]" if result["found"] else "[NOT FOUND]"
    title  = result["title"]
    print(
        f"{status:<{_W_ST}} | "
        f"{str(result['num']):<{_W_NUM}} | "
        f"{title}"
    )

def print_summary(stats: dict, commentary: str = ""):
    print("\n" + "=" * _COL)
    print("FINAL STATISTICS")
    print("=" * _COL)
    print(f"Total Books        : {stats['total']}")
    print(f"Verified Books     : {stats['verified']}")
    print(f"Not Verified       : {stats['not_verified']}")
    print(f"Proportion Found   : {stats['proportion']*100:.2f}%")
    print(f"Margin of Error    : +-{stats['margin_error']*100:.2f}%")
    print(f"95% CI Range       : {stats['lower']*100:.2f}% -> {stats['upper']*100:.2f}%")
    print(f"Avg Confidence     : {stats['avg_confidence']*100:.2f}%")
    print(f"Gemini API Calls   : {_gemini_call_count} / {GEMINI_MAX_CALLS}")
    if commentary:
        print(f"\nAI Summary: {commentary}")


# RUN

def run_checks(books: list[dict]) -> list[dict]:
    results = []
    for book in books:
        result = resolve_title(book)
        results.append(result)
        print_result(result)
        if LIVE_MODE:
            time.sleep(0.3)
    return results


# EXPORT
def export_excel(results: list[dict], stats: dict, commentary: str = ""):
    wb = Workbook()

    ws = wb.active
    ws.title = "Results"

    headers = [
        "#", "Original Title", "Author", "Genre",
        "Status", "Translated Title", "Translation Status",
        "Search Title (ASCII)", "Wikipedia Title",
        "URL", "Confidence", "Resolution Method", "Duplicate Group",
    ]
    col_widths = [5, 35, 25, 15, 12, 35, 18, 35, 35, 50, 12, 22, 15]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, name="Arial")
        ws.column_dimensions[cell.column_letter].width = w

    for i, r in enumerate(results, 2):
        ws.cell(i, 1,  r["num"])
        ws.cell(i, 2,  r["title"])
        ws.cell(i, 3,  r["author"])
        ws.cell(i, 4,  r["genre"])
        ws.cell(i, 5,  "[VERIFIED]" if r["found"] else "[NOT FOUND]")
        ws.cell(i, 6,  r["translated_title"])
        ws.cell(i, 7,  r["translation_status"])
        ws.cell(i, 8,  r["search_title_used"])
        ws.cell(i, 9,  r["wiki_title"])
        ws.cell(i, 10, r["url"])
        ws.cell(i, 11, r["confidence"])
        ws.cell(i, 12, r["resolution_method"])
        ws.cell(i, 13, r["duplicate_group"])

    ws2 = wb.create_sheet("Stats")
    ws2["A1"] = "BOOK VERIFICATION STATISTICS"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")

    summary_rows = [
        ("Total Books",          stats["total"]),
        ("Verified Books",       stats["verified"]),
        ("Not Verified Books",   stats["not_verified"]),
        ("Proportion Verified",  f"{stats['proportion']*100:.2f}%"),
        ("Margin of Error",      f"+-{stats['margin_error']*100:.2f}%"),
        ("Lower Bound (95%)",    f"{stats['lower']*100:.2f}%"),
        ("Upper Bound (95%)",    f"{stats['upper']*100:.2f}%"),
        ("Avg Match Confidence", f"{stats['avg_confidence']*100:.2f}%"),
        ("Gemini Calls Used",    f"{_gemini_call_count} / {GEMINI_MAX_CALLS}"),
    ]
    if commentary:
        summary_rows.append(("AI Summary", commentary))

    for i, (k, v) in enumerate(summary_rows, 3):
        ws2.cell(i, 1, k).font = Font(bold=True, name="Arial")
        ws2.cell(i, 2, v)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 80

    wb.save("books_verified.xlsx")


# MAIN

def main():
    print_header()
    books = load_books()
    print(f"Loaded {len(books)} books\n")

    results = run_checks(books)

    print("\n  [GEMINI] Running duplicate detection...")
    results = gemini_find_duplicates(results)

    stats = compute_confidence_interval(results)

    print("  [GEMINI] Generating stats commentary...")
    commentary = gemini_confidence_commentary(stats)

    print_summary(stats, commentary)
    export_excel(results, stats, commentary)

    print()
    print("Verification Finished!")
    print("File saved as books_verified.xlsx")


if __name__ == "__main__":
    main()